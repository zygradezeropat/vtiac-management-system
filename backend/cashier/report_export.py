"""Excel export for cashier collection reports."""

from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CashierPayment
from .transactions import _local_dt, _parse_date_param, _range_bounds, _particulars_summary

INSTITUTION_NAME = "Valiant Technical Institute Assessment Center, Inc. (VTIAC)"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _payments_queryset():
    return CashierPayment.objects.select_related(
        "profile", "registration", "recorded_by"
    ).order_by("created_at")


def _payments_for_day(day: date):
    start_dt, end_dt = _range_bounds(day, day)
    return list(_payments_queryset().filter(created_at__gte=start_dt, created_at__lte=end_dt))


def _payments_for_month(year: int, month: int):
    last_day = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)
    start_dt, end_dt = _range_bounds(start, end)
    return list(_payments_queryset().filter(created_at__gte=start_dt, created_at__lte=end_dt))


def _payments_for_year(year: int):
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    start_dt, end_dt = _range_bounds(start, end)
    return list(_payments_queryset().filter(created_at__gte=start_dt, created_at__lte=end_dt))


def _program_for_payment(payment: CashierPayment) -> str:
    profile = payment.profile
    if profile and profile.selected_program:
        return profile.selected_program
    if payment.registration_id and payment.registration:
        return payment.registration.selected_program or "—"
    return "—"


def _cashier_name(payment: CashierPayment) -> str:
    user = payment.recorded_by
    if not user:
        return "—"
    return user.get_full_name() or user.get_username() or "—"


def _money(value) -> Decimal:
    return Decimal(str(value or 0))


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, *, bold=False, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if bold:
        cell.font = Font(bold=True)
    if fill:
        cell.fill = fill


def _write_title_block(ws, title: str, meta_rows: list[tuple[str, str]], col_count: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    org_cell = ws.cell(row=2, column=1, value=INSTITUTION_NAME)
    org_cell.font = SUBTITLE_FONT
    org_cell.alignment = Alignment(horizontal="center")

    row = 4
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_count)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _autosize_columns(ws, col_count: int, max_width: int = 42):
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def build_daily_workbook(day: date) -> tuple[BytesIO, str]:
    payments = _payments_for_day(day)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Consolidated"

    headers = [
        "S.No.",
        "Date & Time",
        "Control No.",
        "OR No.",
        "Student Name",
        "Course / Program",
        "Particulars",
        "Total Payable",
        "Amount Paid",
        "Balance",
        "Status",
        "Recorded By",
    ]
    col_count = len(headers)
    header_row = _write_title_block(
        ws,
        "CONSOLIDATED DAILY COLLECTION REPORT",
        [
            ("Report Date", day.strftime("%B %d, %Y")),
            ("Generated On", timezone.localtime(timezone.now()).strftime("%B %d, %Y · %I:%M %p")),
            ("Total Transactions", str(len(payments))),
        ],
        col_count,
    )

    for idx, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    _style_header_row(ws, header_row, col_count)

    total_paid = Decimal("0")
    data_row = header_row + 1
    for i, payment in enumerate(payments, start=1):
        paid = _money(payment.paid_amount)
        total_paid += paid
        values = [
            i,
            _local_dt(payment.created_at).strftime("%b %d, %Y · %I:%M %p"),
            payment.control_number,
            payment.or_number or "—",
            payment.student_name,
            _program_for_payment(payment),
            _particulars_summary(payment.particulars),
            float(_money(payment.total_payable)),
            float(paid),
            float(_money(payment.remaining_balance)),
            payment.status,
            _cashier_name(payment),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=data_row, column=col, value=value)
            _style_data_cell(ws, data_row, col)
        data_row += 1

    if not payments:
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=col_count)
        ws.cell(row=data_row, column=1, value="No transactions recorded for this date.")
        data_row += 1

    total_row = data_row
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    paid_cell = ws.cell(row=total_row, column=9, value=float(total_paid))
    paid_cell.font = TOTAL_FONT
    paid_cell.fill = TOTAL_FILL
    for col in range(1, col_count + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.font = TOTAL_FONT

    _autosize_columns(ws, col_count)
    filename = f"VTIAC_Daily_Collection_{day.isoformat()}.xlsx"
    return _workbook_to_buffer(wb), filename


def build_monthly_workbook(year: int, month: int) -> tuple[BytesIO, str]:
    payments = _payments_for_month(year, month)
    wb = Workbook()

    # Sheet 1 — transaction detail
    ws = wb.active
    ws.title = "Monthly Detail"
    headers = [
        "S.No.",
        "Date",
        "Control No.",
        "Student Name",
        "Course / Program",
        "Amount Paid",
        "Status",
    ]
    col_count = len(headers)
    period_label = date(year, month, 1).strftime("%B %Y")
    header_row = _write_title_block(
        ws,
        "MONTHLY COLLECTION SUMMARY REPORT",
        [
            ("Report Period", period_label),
            ("Generated On", timezone.localtime(timezone.now()).strftime("%B %d, %Y · %I:%M %p")),
            ("Total Transactions", str(len(payments))),
        ],
        col_count,
    )
    for idx, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    _style_header_row(ws, header_row, col_count)

    total_paid = Decimal("0")
    data_row = header_row + 1
    for i, payment in enumerate(payments, start=1):
        paid = _money(payment.paid_amount)
        total_paid += paid
        values = [
            i,
            _local_dt(payment.created_at).strftime("%Y-%m-%d"),
            payment.control_number,
            payment.student_name,
            _program_for_payment(payment),
            float(paid),
            payment.status,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=data_row, column=col, value=value)
            _style_data_cell(ws, data_row, col)
        data_row += 1

    total_row = data_row
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    ws.cell(row=total_row, column=6, value=float(total_paid)).font = TOTAL_FONT
    for col in range(1, col_count + 1):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        ws.cell(row=total_row, column=col).font = TOTAL_FONT

    _autosize_columns(ws, col_count)

    # Sheet 2 — by program + by day
    ws2 = wb.create_sheet("Summary Breakdown")
    summary_headers = ["Category", "Group", "Transactions", "Total Collected (₱)"]
    s_col = len(summary_headers)
    row = _write_title_block(
        ws2,
        "MONTHLY SUMMARY BREAKDOWN",
        [("Report Period", period_label)],
        s_col,
    )
    for idx, label in enumerate(summary_headers, start=1):
        ws2.cell(row=row, column=idx, value=label)
    _style_header_row(ws2, row, s_col)
    row += 1

    by_program: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    by_day: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    for payment in payments:
        program = _program_for_payment(payment)
        day_key = _local_dt(payment.created_at).strftime("%Y-%m-%d")
        paid = _money(payment.paid_amount)
        by_program[program]["count"] += 1
        by_program[program]["amount"] += paid
        by_day[day_key]["count"] += 1
        by_day[day_key]["amount"] += paid

    ws2.cell(row=row, column=1, value="By Course / Program").font = Font(bold=True)
    row += 1
    for program in sorted(by_program.keys()):
        stats = by_program[program]
        ws2.cell(row=row, column=1, value="Program")
        ws2.cell(row=row, column=2, value=program)
        ws2.cell(row=row, column=3, value=stats["count"])
        ws2.cell(row=row, column=4, value=float(stats["amount"]))
        for col in range(1, s_col + 1):
            _style_data_cell(ws2, row, col)
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="By Day").font = Font(bold=True)
    row += 1
    for day_key in sorted(by_day.keys()):
        stats = by_day[day_key]
        ws2.cell(row=row, column=1, value="Date")
        ws2.cell(row=row, column=2, value=day_key)
        ws2.cell(row=row, column=3, value=stats["count"])
        ws2.cell(row=row, column=4, value=float(stats["amount"]))
        for col in range(1, s_col + 1):
            _style_data_cell(ws2, row, col)
        row += 1

    _autosize_columns(ws2, s_col)
    filename = f"VTIAC_Monthly_Collection_{year}_{month:02d}.xlsx"
    return _workbook_to_buffer(wb), filename


def build_annual_workbook(year: int) -> tuple[BytesIO, str]:
    payments = _payments_for_year(year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Performance"

    headers = ["Month", "Transactions", "Total Collected (₱)", "Full Payments", "Partial Payments"]
    col_count = len(headers)
    header_row = _write_title_block(
        ws,
        "ANNUAL COLLECTION PERFORMANCE REPORT",
        [
            ("Report Year", str(year)),
            ("Generated On", timezone.localtime(timezone.now()).strftime("%B %d, %Y · %I:%M %p")),
            ("Total Transactions (Year)", str(len(payments))),
        ],
        col_count,
    )
    for idx, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    _style_header_row(ws, header_row, col_count)

    monthly: dict[int, dict] = {
        m: {"count": 0, "amount": Decimal("0"), "full": 0, "partial": 0} for m in range(1, 13)
    }
    by_program: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

    for payment in payments:
        dt = _local_dt(payment.created_at)
        bucket = monthly[dt.month]
        paid = _money(payment.paid_amount)
        bucket["count"] += 1
        bucket["amount"] += paid
        if payment.status == CashierPayment.Status.FULL:
            bucket["full"] += 1
        elif payment.status == CashierPayment.Status.PARTIAL:
            bucket["partial"] += 1
        by_program[_program_for_payment(payment)] += paid

    data_row = header_row + 1
    year_total = Decimal("0")
    year_count = 0
    for month in range(1, 13):
        stats = monthly[month]
        year_total += stats["amount"]
        year_count += stats["count"]
        values = [
            date(year, month, 1).strftime("%B"),
            stats["count"],
            float(stats["amount"]),
            stats["full"],
            stats["partial"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=data_row, column=col, value=value)
            _style_data_cell(ws, data_row, col)
        data_row += 1

    total_row = data_row
    ws.cell(row=total_row, column=1, value="YEAR TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=2, value=year_count).font = TOTAL_FONT
    ws.cell(row=total_row, column=3, value=float(year_total)).font = TOTAL_FONT
    for col in range(1, col_count + 1):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER
        ws.cell(row=total_row, column=col).font = TOTAL_FONT

    _autosize_columns(ws, col_count)

    ws2 = wb.create_sheet("By Program (Annual)")
    p_headers = ["Course / Program", "Total Collected (₱)", "Share of Year (%)"]
    p_col = len(p_headers)
    row = _write_title_block(ws2, "ANNUAL COLLECTION BY PROGRAM", [("Report Year", str(year))], p_col)
    for idx, label in enumerate(p_headers, start=1):
        ws2.cell(row=row, column=idx, value=label)
    _style_header_row(ws2, row, p_col)
    row += 1

    for program in sorted(by_program.keys(), key=lambda k: by_program[k], reverse=True):
        amount = by_program[program]
        share = float((amount / year_total * 100) if year_total > 0 else 0)
        ws2.cell(row=row, column=1, value=program)
        ws2.cell(row=row, column=2, value=float(amount))
        ws2.cell(row=row, column=3, value=round(share, 2))
        for col in range(1, p_col + 1):
            _style_data_cell(ws2, row, col)
        row += 1

    _autosize_columns(ws2, p_col)
    filename = f"VTIAC_Annual_Collection_{year}.xlsx"
    return _workbook_to_buffer(wb), filename


def build_report_workbook(report_type: str, *, day: date | None = None, year: int | None = None, month: int | None = None):
    report_type = (report_type or "").strip().lower()
    today = timezone.localdate()

    if report_type == "daily":
        target_day = day or today
        return build_daily_workbook(target_day)
    if report_type == "monthly":
        y = year or today.year
        m = month or today.month
        return build_monthly_workbook(y, m)
    if report_type == "annual":
        y = year or today.year
        return build_annual_workbook(y)
    raise ValueError("Invalid report type. Use daily, monthly, or annual.")


def _workbook_to_buffer(wb: Workbook) -> BytesIO:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
